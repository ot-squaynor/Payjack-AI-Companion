# scripts/kb_build.py
# 2026-04-02
"""Purpose: Chunk local KB source documents into retrieval-ready JSONL artifacts."""

from __future__ import annotations

##BRICK 1: Imports, repository bootstrap, and env loading
import argparse
from datetime import datetime, timezone
import json
import logging
from pathlib import Path
import sys

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]

if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from backend.app.env import load_local_env


load_local_env()
logger = logging.getLogger(__name__)


##BRICK 2: CLI parsing and document readers
def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build chunked KB artifacts from local raw docs.")
    parser.add_argument("--input-dir", default="backend/kb/raw_docs")
    parser.add_argument("--output-dir", default="backend/kb/processed_docs")
    parser.add_argument("--metadata-dir", default="backend/kb/metadata")
    parser.add_argument("--chunk-size", type=int, default=1000)
    parser.add_argument("--chunk-overlap", type=int, default=200)
    parser.add_argument("--log-level", default="INFO")
    return parser.parse_args()


def _read_xlsx(path: Path) -> str:
    workbook = load_workbook(path, data_only=True, read_only=True)
    text_parts: list[str] = []
    for sheet in workbook.worksheets:
        text_parts.append(f"# Sheet: {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            values = [str(value).strip() for value in row if value is not None and str(value).strip()]
            if values:
                text_parts.append(" | ".join(values))
    return "\n".join(text_parts)


def _read_text_document(path: Path) -> str:
    if path.suffix.lower() in {".md", ".txt"}:
        return path.read_text(encoding="utf-8", errors="replace")
    if path.suffix.lower() == ".json":
        payload = json.loads(path.read_text(encoding="utf-8"))
        return json.dumps(payload, indent=2, sort_keys=True)
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        return _read_xlsx(path)
    raise ValueError(f"Unsupported KB document format: {path.suffix}")


##BRICK 3: Chunk construction helpers
def _split_into_chunks(text: str, *, chunk_size: int, chunk_overlap: int) -> list[str]:
    paragraphs = [paragraph.strip() for paragraph in text.splitlines() if paragraph.strip()]
    normalized_text = "\n".join(paragraphs).strip()
    if not normalized_text:
        return []

    safe_overlap = max(0, min(chunk_overlap, max(chunk_size - 1, 0)))
    chunks: list[str] = []
    start = 0
    text_length = len(normalized_text)

    while start < text_length:
        target_end = min(text_length, start + chunk_size)
        end = target_end

        if target_end < text_length:
            paragraph_boundary = normalized_text.rfind("\n", start, target_end)
            word_boundary = normalized_text.rfind(" ", start, target_end)
            if paragraph_boundary > start + (chunk_size // 2):
                end = paragraph_boundary
            elif word_boundary > start + (chunk_size // 2):
                end = word_boundary

        chunk_text = normalized_text[start:end].strip()
        if chunk_text and (not chunks or chunks[-1] != chunk_text):
            chunks.append(chunk_text)

        if end >= text_length:
            break

        next_start = max(end - safe_overlap, start + 1)
        while next_start < text_length and normalized_text[next_start].isspace():
            next_start += 1
        start = next_start

    return chunks


##BRICK 4: CLI entrypoint
def main() -> None:
    load_local_env()
    args = parse_args()
    logging.basicConfig(
        level=getattr(logging, args.log_level.upper(), logging.INFO),
        format="%(asctime)s %(levelname)s %(name)s %(message)s",
    )

    input_dir = Path(args.input_dir)
    output_dir = Path(args.output_dir)
    metadata_dir = Path(args.metadata_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    metadata_dir.mkdir(parents=True, exist_ok=True)

    chunk_records: list[dict[str, object]] = []
    processed_documents: list[str] = []

    supported_suffixes = {".md", ".txt", ".json", ".xlsx", ".xlsm"}
    for path in sorted(input_dir.rglob("*")):
        if not path.is_file() or path.suffix.lower() not in supported_suffixes:
            continue
        relative_path = path.relative_to(input_dir).as_posix()
        doc_id = relative_path.replace("/", "_").replace(".", "_")
        title = path.stem.replace("_", " ").replace("-", " ").title()
        try:
            text = _read_text_document(path)
        except Exception as exc:  # pragma: no cover - depends on local document formats
            logger.warning("Skipping KB source %s: %s", relative_path, exc)
            continue
        if not text.strip():
            logger.warning("Skipping empty KB source %s", relative_path)
            continue
        chunks = _split_into_chunks(
            text,
            chunk_size=args.chunk_size,
            chunk_overlap=args.chunk_overlap,
        )
        processed_documents.append(relative_path)
        for chunk_index, chunk_text in enumerate(chunks):
            chunk_records.append(
                {
                    "doc_id": doc_id,
                    "chunk_id": chunk_index,
                    "title": title,
                    "text": chunk_text,
                    "metadata": {
                        "source_path": relative_path,
                        "type": path.parts[-2] if len(path.parts) >= 2 else "general",
                    },
                }
            )

    chunks_path = output_dir / "chunks.jsonl"
    with chunks_path.open("w", encoding="utf-8") as handle:
        for record in chunk_records:
            handle.write(json.dumps(record, ensure_ascii=False) + "\n")

    manifest = {
        "build_id": datetime.now(timezone.utc).strftime("kb_%Y%m%dT%H%M%SZ"),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "documents": processed_documents,
        "chunk_count": len(chunk_records),
    }
    (metadata_dir / "kb_manifest.json").write_text(
        json.dumps(manifest, indent=2, sort_keys=True),
        encoding="utf-8",
    )

    logger.info("KB build complete. documents=%s chunks=%s", len(processed_documents), len(chunk_records))


if __name__ == "__main__":
    main()
